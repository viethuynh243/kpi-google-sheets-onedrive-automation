"""
automate_kpi.py  –  Google Sheets → Vốn hóa Chi phí Nhân sự 2026
Tải ACCA, CMA, IT từ Google Sheets → trích xuất → ghi vào file Excel template

Usage:
    python automate_kpi.py                    # dùng file local đã tải sẵn
    python automate_kpi.py --download         # tải mới từ Google Sheets rồi xử lý
    python automate_kpi.py --download --open  # tải + mở file kết quả tự động
"""

from __future__ import annotations

import argparse
import io
import re
import shutil
import sys
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Force UTF-8 stdout so Vietnamese text prints correctly on Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG – chỉnh ID nếu Google Sheet đổi link
# ─────────────────────────────────────────────────────────────────────────────

WORK_DIR = Path(__file__).parent

DIRS = {
    "raw":      WORK_DIR / "data" / "input" / "raw",
    "template": WORK_DIR / "data" / "input" / "template",
    "staging":  WORK_DIR / "data" / "output" / "staging",
    "final":    WORK_DIR / "data" / "output" / "final",
    "logs":     WORK_DIR / "logs",
}

SOURCES = {
    "CMA":  {
        "id":         "1jOaBolZ78dbelYkoFL5jSUE0-yJn5425",
        "file":       "CMA.xlsx",
        "department": "SX ACCA+CMA",
        "type":       "kpi",
    },
    "ACCA": {
        "id":         "16w4-UpSFnjVGpMJlfY9m8LIPTdMZy1dQ",
        "file":       "ACCA.xlsx",
        "department": "SX ACCA+CMA",
        "type":       "kpi",
    },
    "IT":   {
        "id":         "1x9FBjRHISImjCII7GqOcTTn40_BmAnRN",
        "file":       "IT.xlsx",
        "department": "IT",
        "type":       "it_matrix",
    },
}

TEMPLATE_FILE    = "von_hoa_template.xlsx"
SHEET_DATA_SX    = "Data SX ACCA+CMA"
SHEET_DATA_CFA   = "Data SX CFA"
SHEET_IT         = "Timesheet IT"

DATA_CFA_START_ROW  = 4   # Hàng đầu tiên ghi data trong sheet Data SX CFA
STANDARD_HOURS_FT   = 176.0   # Giờ/tháng tiêu chuẩn cho FT employee
STANDARD_HOURS_PT   = 80.0    # Giờ/tháng tiêu chuẩn cho PT employee

# Thứ tự cột khi ghi vào sheet "Data SX ACCA+CMA" (1-indexed col 1..18)
# None = bỏ trống (MNV không có trong dữ liệu nguồn)
COLUMN_ORDER = [
    "month",             # col 1  – Tháng
    "program",           # col 2  – BU / Chương trình
    "position",          # col 3  – Vị trí
    "employee",          # col 4  – Tên NV
    None,                # col 5  – MNV (bỏ trống)
    "product_or_project",# col 6  – Tên sản phẩm
    "reference_link",    # col 7  – Link
    "new_or_old",        # col 8  – SP mới/cũ
    "project_name",      # col 9  – Tên dự án
    "subject_or_system", # col 10 – Bộ môn
    "product_feature",   # col 11 – Đặc tính SP
    "deliverable",       # col 12 – SP bàn giao
    "component",         # col 13 – Cấu phần
    "complexity",        # col 14 – Độ khó
    "unit",              # col 15 – Đơn vị tính
    "actual_quantity",   # col 16 – Số lượng actual
    "kpi_standard",      # col 17 – KPI standard
    "total_kpi",         # col 18 – Total KPI (h)
]

DATA_SX_START_ROW = 4   # Hàng bắt đầu ghi data (sau 3 hàng header)
IT_HEADER_ROW     = 13  # Hàng đầu tiên cần cập nhật trong Timesheet IT


# ─────────────────────────────────────────────────────────────────────────────
#  UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        v = re.sub(r"\s+", " ", value).strip()
        return v or None
    return value


def norm(value: Any) -> str:
    """Chuẩn hóa header: lowercase + collapse whitespace."""
    t = clean(value)
    return str(t).lower() if t is not None else ""


def is_num(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def infer_month_year(sheet_name: str) -> tuple[int | None, int | None]:
    """Suy ra tháng/năm từ tên sheet (vd: 'Apr 26', 'T4 2026')."""
    text = sheet_name.lower().strip()
    months = {
        "jan": 1, "january": 1, "feb": 2, "february": 2,
        "mar": 3, "march": 3, "apr": 4, "april": 4, "may": 5,
        "jun": 6, "june": 6, "jul": 7, "july": 7,
        "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10, "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    year_m = re.search(r"\b(20\d{2}|\d{2})\b", text)
    year = None
    if year_m:
        y = int(year_m.group(1))
        year = y if y > 100 else 2000 + y
    month = None
    for name, num in months.items():
        if re.search(rf"\b{name}\b", text):
            month = num
            break
    if month is None:
        m2 = re.search(r"\b(?:t[àha]?[àng]*|month|thang)\s*(\d{1,2})\b", text)
        if m2:
            month = int(m2.group(1))
    return year, month


# ─────────────────────────────────────────────────────────────────────────────
#  SETUP
# ─────────────────────────────────────────────────────────────────────────────

def setup_dirs() -> None:
    for d in DIRS.values():
        d.mkdir(parents=True, exist_ok=True)


def find_template() -> Path | None:
    """Tìm file template trong các vị trí thông thường."""
    candidates = [
        DIRS["template"] / TEMPLATE_FILE,
        WORK_DIR / "template" / TEMPLATE_FILE,
    ]
    # Cũng tìm file .xlsx nào chứa 'von_hoa' hoặc 'huong_dan' trong tên
    for d in [DIRS["template"], WORK_DIR / "template"]:
        if d.exists():
            for f in d.glob("*.xlsx"):
                if any(x in f.name.lower() for x in ["von_hoa", "huong_dan"]):
                    candidates.insert(0, f)
    for p in candidates:
        if p.exists():
            return p
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────

def download_all() -> None:
    for key, meta in SOURCES.items():
        url = f"https://docs.google.com/spreadsheets/d/{meta['id']}/export?format=xlsx"
        out = DIRS["raw"] / meta["file"]
        tmp = DIRS["raw"] / f".{out.stem}.download.xlsx"
        print(f"  {key}...", end=" ", flush=True)
        try:
            urllib.request.urlretrieve(url, tmp)
            try:
                tmp.replace(out)
            except PermissionError:
                shutil.copy2(tmp, out)
                tmp.unlink(missing_ok=True)
            print(f"OK  ({out.stat().st_size // 1024} KB)")
        except Exception as e:
            print(f"FAILED: {e}")
            raise


# ─────────────────────────────────────────────────────────────────────────────
#  EXTRACT: ACCA + CMA (KPI long-format)
# ─────────────────────────────────────────────────────────────────────────────

def _find_header_row(ws, max_row: int = 40) -> tuple[int | None, dict[str, int]]:
    """Tìm hàng header: hàng chứa cả 'tên nhân viên' và ('tên sản phẩm' hoặc 'năm')."""
    for row_idx in range(1, min(ws.max_row or 0, max_row) + 1):
        hdrs: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            h = norm(ws.cell(row_idx, col).value)
            if h:
                hdrs[h] = col
        if "tên nhân viên" in hdrs and ("tên sản phẩm" in hdrs or "năm" in hdrs):
            return row_idx, hdrs
    return None, {}


def _val(ws, row: int, hdrs: dict[str, int], *keys: str) -> Any:
    """Lấy giá trị theo tên cột (hỗ trợ contains match khi exact match thất bại)."""
    for key in keys:
        key_l = key.lower()
        # 1. Exact match
        col = hdrs.get(key_l)
        if col:
            return clean(ws.cell(row, col).value)
        # 2. Contains match (dùng cho header có text thừa như newline, ngoặc)
        for h, c in hdrs.items():
            if key_l in h:
                return clean(ws.cell(row, c).value)
    return None


def extract_kpi_rows(path: Path, department: str) -> list[dict[str, Any]]:
    """Trích xuất dữ liệu KPI từ ACCA hoặc CMA xlsx (mỗi sheet = 1 nhân viên hoặc 1 tháng)."""
    wb = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        header_row, hdrs = _find_header_row(ws)
        if not header_row:
            continue

        inf_year, inf_month = infer_month_year(ws.title)
        has_ym = "năm" in hdrs and "tháng" in hdrs

        for r in range(header_row + 1, (ws.max_row or 0) + 1):
            employee = _val(ws, r, hdrs, "tên nhân viên")
            product  = _val(ws, r, hdrs, "tên sản phẩm")
            actual   = _val(ws, r, hdrs, "số lượng actual")
            kpi_std  = _val(ws, r, hdrs, "kpi standard")
            total    = _val(ws, r, hdrs, "total kpi")

            # Bỏ qua hàng rỗng
            if not any([employee, product, actual]):
                continue

            # Tính total_kpi nếu thiếu
            if total is None and is_num(actual) and is_num(kpi_std):
                total = round(float(actual) * float(kpi_std), 4)

            year  = _val(ws, r, hdrs, "năm")   if has_ym else inf_year
            month = _val(ws, r, hdrs, "tháng") if has_ym else inf_month

            rows.append({
                "source_file":        path.name,
                "source_sheet":       ws.title,
                "year":               year,
                "month":              month,
                "department":         department,
                "program":            _val(ws, r, hdrs, "chương trình"),
                "position":           _val(ws, r, hdrs, "vị trí"),
                "employee":           employee,
                "product_or_project": product,
                "reference_link":     _val(ws, r, hdrs,
                                          "link tham chiếu",
                                          "link"),
                "new_or_old":         _val(ws, r, hdrs,
                                          "sản phẩm mới/sản phẩm cũ",
                                          "sản phẩm mới/cũ"),
                "project_name":       _val(ws, r, hdrs, "tên dự án"),
                "subject_or_system":  _val(ws, r, hdrs, "bộ môn"),
                "product_feature":    _val(ws, r, hdrs,
                                          "đặc tính sản phẩm",
                                          "đặc tính sp",
                                          "đặc tính"),
                "deliverable":        _val(ws, r, hdrs, "sản phẩm bàn giao"),
                "component":          _val(ws, r, hdrs, "cấu phần"),
                "complexity":         _val(ws, r, hdrs, "độ khó"),
                "unit":               _val(ws, r, hdrs, "đơn vị tính"),
                "actual_quantity":    actual,
                "kpi_standard":       kpi_std,
                "total_kpi":          total,
            })

    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  EXTRACT: IT (matrix – giữ nguyên dạng để copy trực tiếp)
# ─────────────────────────────────────────────────────────────────────────────

def read_it_sheet(path: Path) -> tuple[str, list[list[Any]]]:
    """Đọc sheet IT năm gần nhất, trả về (sheet_name, list_of_rows)."""
    wb = load_workbook(path, data_only=True)
    # Ưu tiên sheet tên là năm số (2026, 2025...)
    target_ws = None
    for ws in sorted(wb.worksheets, key=lambda w: w.title, reverse=True):
        if str(ws.title).isdigit():
            target_ws = ws
            break
    if target_ws is None:
        target_ws = wb.worksheets[0]

    rows = [
        [ws.cell(r, c).value for c in range(1, (target_ws.max_column or 0) + 1)]
        for r in range(1, (target_ws.max_row or 0) + 1)
    ]
    return target_ws.title, rows


# ─────────────────────────────────────────────────────────────────────────────
#  WRITE: Data SX ACCA+CMA
# ─────────────────────────────────────────────────────────────────────────────

def write_data_sx(ws, rows: list[dict[str, Any]]) -> None:
    # Xóa data cũ từ hàng DATA_SX_START_ROW trở xuống
    for r in range(DATA_SX_START_ROW, ws.max_row + 1):
        for c in range(1, len(COLUMN_ORDER) + 1):
            ws.cell(r, c).value = None

    for offset, row in enumerate(rows):
        r = DATA_SX_START_ROW + offset
        for c_idx, field in enumerate(COLUMN_ORDER, start=1):
            ws.cell(r, c_idx).value = row.get(field) if field else None


# ─────────────────────────────────────────────────────────────────────────────
#  WRITE: Data SX CFA (capacity rate matrix theo nhân viên × tháng)
# ─────────────────────────────────────────────────────────────────────────────

def _build_cfa_matrix(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Gom KPI rows theo nhân viên → tính capacity rate cho từng tháng."""
    from collections import defaultdict

    emp_months: dict[str, dict] = {}
    for row in rows:
        emp = row.get("employee")
        if not emp:
            continue
        month_val = row.get("month")
        if not is_num(month_val):
            continue
        m = int(float(month_val))
        if not (1 <= m <= 12):
            continue
        total = row.get("total_kpi")
        if not is_num(total):
            continue

        if emp not in emp_months:
            emp_months[emp] = {
                "position": row.get("position"),
                "project":  row.get("project_name") or row.get("program") or "CFA",
                "months":   defaultdict(float),
            }
        emp_months[emp]["months"][m] += float(total)

    result = []
    for emp, data in sorted(emp_months.items()):
        pos = (data["position"] or "").lower()
        std = STANDARD_HOURS_PT if ("part-time" in pos or pos == "pt") else STANDARD_HOURS_FT
        capacity = {m: round(data["months"].get(m, 0.0) / std, 4) for m in range(1, 13)}
        result.append({
            "project":  data["project"],
            "employee": emp,
            "capacity": capacity,
        })
    return result


def write_data_cfa(ws, rows: list[dict[str, Any]]) -> int:
    """Ghi dữ liệu vào sheet 'Data SX CFA'. Trả về số nhân viên đã ghi."""
    matrix = _build_cfa_matrix(rows)
    if not matrix:
        return 0

    # Xóa data cũ từ DATA_CFA_START_ROW trở xuống
    for r in range(DATA_CFA_START_ROW, ws.max_row + 1):
        for c in range(1, 17):
            try:
                ws.cell(r, c).value = None
            except AttributeError:
                pass

    for idx, emp_data in enumerate(matrix, start=1):
        r = DATA_CFA_START_ROW + idx - 1
        ws.cell(r, 1).value = idx                       # TT
        ws.cell(r, 2).value = emp_data["project"]       # Tên dự án
        ws.cell(r, 3).value = emp_data["employee"]      # Tên nhân viên
        ws.cell(r, 4).value = None                      # MNV (không có trong source)
        for m in range(1, 13):
            ws.cell(r, 4 + m).value = emp_data["capacity"][m]  # cols 5..16

    return len(matrix)


# ─────────────────────────────────────────────────────────────────────────────
#  WRITE: Timesheet IT (copy matrix trực tiếp từ nguồn)
# ─────────────────────────────────────────────────────────────────────────────

def write_it_sheet(ws, it_rows: list[list[Any]]) -> None:
    """
    Ghi dữ liệu IT dạng matrix vào sheet Timesheet IT.
    Chỉ cập nhật từ hàng IT_HEADER_ROW trở xuống.
    it_rows: list of rows tính từ 1-indexed row 1 của sheet nguồn IT.
    """
    # Unmerge tất cả merged cells trong vùng cần ghi để tránh AttributeError
    merges_to_remove = [
        mc for mc in list(ws.merged_cells.ranges)
        if mc.min_row >= IT_HEADER_ROW
    ]
    for mc in merges_to_remove:
        ws.unmerge_cells(str(mc))

    # Xóa data cũ từ IT_HEADER_ROW trở xuống
    for r in range(IT_HEADER_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            try:
                ws.cell(r, c).value = None
            except AttributeError:
                pass  # bỏ qua merged cell còn sót

    # Copy từ hàng IT_HEADER_ROW - 1 (0-indexed) của it_rows
    for offset, row in enumerate(it_rows[IT_HEADER_ROW - 1:]):
        r = IT_HEADER_ROW + offset
        for c_idx, value in enumerate(row, start=1):
            try:
                ws.cell(r, c_idx).value = value
            except AttributeError:
                pass


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run(download: bool, open_after: bool = False) -> Path:
    setup_dirs()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"\n{'='*60}")
    print(f"Vốn hóa Chi phí Nhân sự 2026  –  {datetime.now():%d/%m/%Y %H:%M}")
    print(f"{'='*60}")

    # 1. Download
    if download:
        print("\n[1/4] Đang tải từ Google Sheets...")
        download_all()
    else:
        print("\n[1/4] Dùng file local (thêm --download để tải mới)")

    # 2. Kiểm tra file raw
    missing = [k for k, m in SOURCES.items()
               if not (DIRS["raw"] / m["file"]).exists()]
    if missing:
        raise SystemExit(
            f"\nKhông tìm thấy file: {missing}\n"
            f"Chạy lại với --download để tải từ Google Sheets."
        )

    # 3. Extract
    print("\n[2/4] Trích xuất dữ liệu...")
    sx_rows: list[dict] = []
    for key in ("CMA", "ACCA"):
        meta = SOURCES[key]
        path = DIRS["raw"] / meta["file"]
        extracted = extract_kpi_rows(path, meta["department"])
        print(f"  {key:<6} → {len(extracted):>5} hàng")
        sx_rows.extend(extracted)

    total_kpi_filled = sum(1 for r in sx_rows if r.get("total_kpi") is not None)
    print(f"  {'SX total':<6} → {len(sx_rows):>5} hàng  "
          f"(Total KPI có giá trị: {total_kpi_filled})")

    it_sheet_name, it_rows = read_it_sheet(DIRS["raw"] / SOURCES["IT"]["file"])
    print(f"  {'IT':<6} → {len(it_rows):>5} hàng  (sheet '{it_sheet_name}')")

    # 4. Tìm / copy template
    print("\n[3/4] Chuẩn bị template...")
    tpl = find_template()
    if tpl is None:
        # Thử tìm trong thư mục template gốc
        root_tpl = WORK_DIR / "template"
        if root_tpl.exists():
            xlsxs = list(root_tpl.glob("*.xlsx"))
            if xlsxs:
                tpl = xlsxs[0]
    if tpl is None:
        raise SystemExit(
            f"\nKhông tìm thấy file template.\n"
            f"Copy file Excel template vào: {DIRS['template']}\n"
            f"Đặt tên là: {TEMPLATE_FILE}"
        )
    print(f"  Template: {tpl.name}")

    # 5. Ghi vào output
    print("\n[4/4] Ghi vào file kết quả...")
    output_path = DIRS["final"] / f"von_hoa_{ts}.xlsx"
    shutil.copy2(tpl, output_path)
    wb = load_workbook(output_path)

    written_sx = False
    written_cfa = False
    written_it = False

    if SHEET_DATA_SX in wb.sheetnames:
        acca_cma_rows = [r for r in sx_rows if r.get("program") in ("ACCA", "CMA")]
        write_data_sx(wb[SHEET_DATA_SX], acca_cma_rows)
        written_sx = True
        print(f"  ✓ '{SHEET_DATA_SX}' ← {len(acca_cma_rows)} hàng (ACCA+CMA)")
    else:
        print(f"  ✗ Sheet '{SHEET_DATA_SX}' không có trong template")

    if SHEET_DATA_CFA in wb.sheetnames:
        cfa_rows = [r for r in sx_rows if r.get("program") == "CFA"]
        n_emp = write_data_cfa(wb[SHEET_DATA_CFA], cfa_rows)
        if n_emp:
            written_cfa = True
            print(f"  ✓ '{SHEET_DATA_CFA}' ← {len(cfa_rows)} KPI rows → {n_emp} nhân viên (capacity matrix)")
        else:
            print(f"  ~ '{SHEET_DATA_CFA}' không có dữ liệu CFA để ghi")
    else:
        print(f"  ✗ Sheet '{SHEET_DATA_CFA}' không có trong template")

    if SHEET_IT in wb.sheetnames:
        write_it_sheet(wb[SHEET_IT], it_rows)
        written_it = True
        print(f"  ✓ '{SHEET_IT}' ← {len(it_rows) - IT_HEADER_ROW + 1} hàng (từ hàng {IT_HEADER_ROW})")
    else:
        print(f"  ✗ Sheet '{SHEET_IT}' không có trong template")

    wb.save(output_path)

    # 6. Lưu staging CSV (để debug / kiểm tra)
    try:
        import csv
        csv_path = DIRS["staging"] / f"normalized_{ts}.csv"
        if sx_rows:
            with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=list(sx_rows[0].keys()))
                writer.writeheader()
                writer.writerows(sx_rows)
    except Exception:
        pass

    # 7. Kết quả
    acca_cma_count = sum(1 for r in sx_rows if r.get("program") in ("ACCA", "CMA"))
    cfa_count = sum(1 for r in sx_rows if r.get("program") == "CFA")
    print(f"\n{'='*60}")
    print(f"HOÀN THÀNH!")
    print(f"  File output   : {output_path}")
    print(f"  SX ACCA+CMA   : {acca_cma_count} hàng  (written={written_sx})")
    print(f"  Data SX CFA   : {cfa_count} KPI rows  (written={written_cfa})")
    print(f"  IT matrix     : sheet={it_sheet_name}, written={written_it}")
    print(f"{'='*60}\n")

    if open_after:
        import subprocess
        subprocess.Popen(["explorer", str(output_path)])

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Vốn hóa Chi phí Nhân sự 2026 – Automation Pipeline"
    )
    parser.add_argument(
        "--download", action="store_true",
        help="Tải mới từ Google Sheets trước khi xử lý"
    )
    parser.add_argument(
        "--open", action="store_true", dest="open_after",
        help="Mở file kết quả trong Excel sau khi hoàn thành"
    )
    args = parser.parse_args()
    run(download=args.download, open_after=args.open_after)


if __name__ == "__main__":
    main()
